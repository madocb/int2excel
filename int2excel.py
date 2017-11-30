#
# int2excel.py
# Script to return interface number, physical and logical up/down status and write to an excel file.
# Populate device-list.txt with host names.

import os.path
import re, getpass, netmiko
from netmiko import ConnectHandler
import openpyxl

# Define pointers
totalpeer = 0
alldownpeer2 = ""
command1 = "show interfaces detail | match physical"
device_type = "juniper"
j_routers = []

scriptpath = os.path.dirname(__file__)
filename = os.path.join(scriptpath, 'device-list.txt')
hostfile = open(filename)

ssh_exceptions = (netmiko.ssh_exception.NetMikoAuthenticationException,
                  netmiko.ssh_exception.NetMikoTimeoutException, netmiko.NetMikoTimeoutException,
                  netmiko.NetMikoAuthenticationException, netmiko.NetmikoTimeoutError, netmiko.NetmikoAuthError,
                  netmiko.ssh_exception.SSHException, netmiko.ssh_exception.AuthenticationException)

un = input('Username: ')
pw = getpass.getpass()
my_file_object = open("device-list.txt", "r")

# Read from hostfile
for line in hostfile:
    if "#" not in line:
        j_routers.append(line.strip())

hostfile.close()

# Below there be Welsh Dragons

def setupexcel():
    wb = openpyxl.Workbook()
    for j_rtr in j_routers:
        wb.create_sheet(title=j_rtr)
        sheet = wb.get_sheet_by_name(j_rtr) 
        sheet['A1'] = 'Interface' 
        sheet['B1'] = 'Logical state'
        sheet['C1'] = 'Physical state' 
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save('int2excel.xlsx')

def write2excel(xa,xb,xc,j_rtr,excelrow):
     wb = openpyxl.load_workbook('int2excel.xlsx')
     sheet = wb.get_sheet_by_name(j_rtr)
     intcell = "A"+str(excelrow)
     logcell = "B"+str(excelrow)
     phycell = "C"+str(excelrow)
     sheet[intcell] = xa
     sheet[logcell] = xb
     sheet[phycell] = xc
     wb.save('int2excel.xlsx')    

def devicetype():
    madshow = ssh_conn.send_command_expect("show version")
    if "cisco" in madshow.lower():
        return "CISCO"
    elif "junos" in madshow.lower():
        return "JUNIPER"
    else:
        return None

setupexcel()

print("Lets Go....:")

for j_rtr in j_routers:
    try:
        print("#" * 79)
        print("Connecting to:", j_rtr)
        ssh_conn = ConnectHandler(ip=j_rtr, device_type=device_type, username=un, password=pw)
        output = ssh_conn.send_command_expect(command1)
        print(devicetype())
        lineoutput = output.splitlines()
        print("Connected to:", j_rtr)

        excelrow = 2
        
        for line in lineoutput:
            if ("xe-" in line or "ge-" in line or "ae" in line):
                totalpeer += 1
                excelrow += 1
                xa =(line.split(" ")[2])
                xb =(line.split(" ")[3])
                xc =(line.split(" ")[-1])
                #print (xa,xb,xc,excelrow)
                write2excel(xa,xb,xc,j_rtr,excelrow)

        ssh_conn.disconnect()
    except ssh_exceptions:
        print("Could not connect to device:", j_rtr)

print("\n")
print("#" * 79)
print("Total number of interfaces collated:", totalpeer)
print("#" * 79)
